from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule


def format_table(worksheet):
    """
    This function gives to the table a nice format with alternating row colors and a header style.
    """
    # Rango completo de la tabla
    ref = worksheet.dimensions  # ejemplo: A1:D100
    
    table = Table(displayName=f"Table_{worksheet.title.replace(' ', '_')}", ref=ref)

    style = TableStyleInfo(
        name="TableStyleMedium9",  # estilo bonito de Excel para título
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,      # filas alternadas en blanco y negro
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    worksheet.add_table(table)

def apply_color_scale(worksheet, column_letter, green_min = 10, yellow_min = 5):
    """ 
    This function applies a color scale to the cells of the 
    specified column based on the quantity values.
    Applies green for values >= green_min, yellow for values >= yellow_min, and red for values < yellow_min.
    """
    green = PatternFill(start_color="66FF33", end_color="66FF33", fill_type="solid")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    if green_min is False:
        green_min = None
        
    if yellow_min is False:
        yellow_min = None
        

    for row in range(2, worksheet.max_row + 1):
        cell = worksheet[f"{column_letter}{row}"]
        
        if cell.value is None:
            continue
        if not isinstance(cell.value, (int, float)):
            continue
        
        if green_min is not None and cell.value >= green_min:
            cell.fill = green
        elif yellow_min is not None and cell.value >= yellow_min:
            cell.fill = yellow
        else:
            cell.fill = red

def apply_gradient_column(
    worksheet,
    column_letter,
    start_color="FFFF00",  # mínimo → amarillo
    end_color="66FF33",    # máximo → verde
    start_row=2
):
    
    max_row = worksheet.max_row

    if max_row < start_row:
        return

    cell_range = f"{column_letter}{start_row}:{column_letter}{max_row}"

    rule = ColorScaleRule(
        start_type='min',
        start_color=start_color,  # valores bajos
        end_type='max',
        end_color=end_color       # valores altos
    )

    worksheet.conditional_formatting.add(cell_range, rule)